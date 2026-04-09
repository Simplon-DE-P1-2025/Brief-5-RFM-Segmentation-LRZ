"""Concatene les feuilles de online_retail_II.xlsx en un seul CSV."""

import csv
from openpyxl import load_workbook

SRC = "online_retail_II.xlsx"
DST = "online_retail_combined.csv"


def main() -> None:
    wb = load_workbook(SRC, read_only=True, data_only=True)
    try:
        with open(DST, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            header_written = False
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = ws.iter_rows(values_only=True)
                header = next(rows)
                if not header_written:
                    writer.writerow(header)
                    header_written = True
                for row in rows:
                    writer.writerow(row)
    finally:
        wb.close()
    print(f"OK : {DST}")


if __name__ == "__main__":
    main()

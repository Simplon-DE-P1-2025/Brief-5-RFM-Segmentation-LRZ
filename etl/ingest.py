"""
etl/ingest.py — Étape 01 du pipeline RFM : ingestion XLSX → raw.online_retail
==============================================================================

Lit `data/online_retail_II.xlsx` en streaming (mémoire constante via openpyxl
read_only), écrit dans un CSV temporaire, puis charge dans Postgres via
COPY FROM STDIN. Idempotent : TRUNCATE optionnel en début d'exécution.

**Note** : ce fichier ne porte PAS le préfixe `01_` car les modules Python
ne peuvent pas commencer par un chiffre. L'ordre d'exécution est :

    01.  etl/ingest.py            (ce fichier)
    02.  etl/02_clean.sql         (cascade exclusive vers clean.*)
    03.  etl/03_rfm.sql           (NTILE(5) → analytics.customer_rfm)
    04.  etl/04_segments.sql      (UPDATE CASE des 11 segments)
    05.  etl/05_view_rfm_v.sql    (vues d'enrichissement)
    06.  (futur) etl/06_history.sql  (snapshot mensuel pour les pages 2/3)

Appelable de deux manières :

  • Standalone (test/fallback) :
        uv run python -m etl.ingest

  • Depuis Airflow PythonOperator :
        from etl.ingest import ingest
        PythonOperator(task_id="ingest_xlsx", python_callable=ingest, ...)

Cette extraction est strictement équivalente — 1 067 371 lignes attendues.
"""

from __future__ import annotations

import csv
import os
import sys
import tempfile
import time
from pathlib import Path

import psycopg2
from openpyxl import load_workbook


# ─────────────────────────────────────────────────────────────────────
# Configuration
# ─────────────────────────────────────────────────────────────────────

ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = Path(os.getenv("DATA_PATH", ROOT / "data" / "online_retail_II.xlsx"))

DSN = os.getenv(
    "RFM_DB_DSN",
    "postgresql://rfm_user:rfm_pass@localhost:5432/rfm_db",
)

SHEETS = ("Year 2009-2010", "Year 2010-2011")


# ─────────────────────────────────────────────────────────────────────
# SQL inline
# ─────────────────────────────────────────────────────────────────────

COPY_RAW_SQL = """
COPY raw.online_retail (invoice, stock_code, description, quantity,
                        invoice_date, price, customer_id, country)
FROM STDIN WITH (FORMAT csv, HEADER true, NULL '')
"""

TRUNCATE_RAW_SQL = "TRUNCATE raw.online_retail RESTART IDENTITY CASCADE;"


# ─────────────────────────────────────────────────────────────────────
# Streaming xlsx → CSV
# ─────────────────────────────────────────────────────────────────────


def xlsx_to_csv(xlsx_path: Path, csv_path: Path) -> int:
    """Streaming des 2 feuilles Excel → 1 CSV. Mémoire constante via openpyxl read_only."""
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Dataset introuvable : {xlsx_path}")

    print(f"[ingest] Lecture streaming de {xlsx_path.name}")
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    n = 0
    try:
        with csv_path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            header_written = False
            for sheet_name in SHEETS:
                ws = wb[sheet_name]
                rows = ws.iter_rows(values_only=True)
                header = next(rows)
                if not header_written:
                    # Header normalisé pour matcher les colonnes du COPY
                    writer.writerow(
                        [
                            "invoice",
                            "stock_code",
                            "description",
                            "quantity",
                            "invoice_date",
                            "price",
                            "customer_id",
                            "country",
                        ]
                    )
                    header_written = True
                else:
                    # On consomme juste l'entête (déjà écrit via la 1re feuille)
                    _ = header
                for row in rows:
                    writer.writerow(row)
                    n += 1
                print(f"[ingest]   feuille « {sheet_name} » → cumul {n:,} lignes")
    finally:
        wb.close()
    return n


# ─────────────────────────────────────────────────────────────────────
# Pipeline d'ingestion
# ─────────────────────────────────────────────────────────────────────


def ingest(
    dsn: str = DSN,
    xlsx_path: Path | str = XLSX_PATH,
    truncate: bool = True,
) -> dict:
    """
    Pipeline complet d'ingestion XLSX → raw.online_retail.

    Étapes :
      0. (optionnel) TRUNCATE raw.online_retail
      1. xlsx → CSV temporaire (streaming, mémoire constante)
      2. COPY FROM STDIN → raw.online_retail
      3. Vérification : count + pg_typeof(customer_id)

    Compatible PythonOperator Airflow : retourne un dict (push XCom natif).
    """
    t0 = time.perf_counter()
    xlsx_path = Path(xlsx_path)

    conn = psycopg2.connect(dsn)
    try:
        if truncate:
            with conn.cursor() as cur:
                cur.execute(TRUNCATE_RAW_SQL)
            conn.commit()
            print("[ingest] TRUNCATE raw.online_retail OK")

        with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as tmp:
            csv_path = Path(tmp.name)

        try:
            n_csv = xlsx_to_csv(xlsx_path, csv_path)
            csv_size_mb = csv_path.stat().st_size / 1e6
            print(f"[ingest] CSV temporaire : {n_csv:,} lignes ({csv_size_mb:.1f} MB)")

            with conn.cursor() as cur, csv_path.open("r", encoding="utf-8") as f:
                cur.copy_expert(COPY_RAW_SQL, f)
            conn.commit()
            print(f"[ingest] COPY → raw.online_retail OK")
        finally:
            csv_path.unlink(missing_ok=True)

        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM raw.online_retail;")
            n_rows = cur.fetchone()[0]
            cur.execute(
                "SELECT pg_typeof(customer_id)::text FROM raw.online_retail LIMIT 1;"
            )
            customer_id_type = cur.fetchone()[0]

        elapsed = round(time.perf_counter() - t0, 1)
        print(
            f"[ingest] Terminé : {n_rows:,} lignes, "
            f"customer_id={customer_id_type}, {elapsed}s"
        )

        return {
            "rows_inserted": n_rows,
            "csv_lines": n_csv,
            "csv_size_mb": round(csv_size_mb, 1),
            "customer_id_type": customer_id_type,
            "elapsed_s": elapsed,
        }
    finally:
        conn.close()


# ─────────────────────────────────────────────────────────────────────
# Entry point standalone
# ─────────────────────────────────────────────────────────────────────


def main() -> int:
    print(f"[ingest] DSN : {DSN}")
    print(f"[ingest] XLSX: {XLSX_PATH}")
    try:
        result = ingest()
    except FileNotFoundError as exc:
        print(f"[ingest] ERREUR : {exc}", file=sys.stderr)
        return 1
    except psycopg2.OperationalError as exc:
        print(f"[ingest] ERREUR connexion Postgres : {exc}", file=sys.stderr)
        print("[ingest] Vérifiez : docker compose up -d postgres", file=sys.stderr)
        return 1
    except Exception as exc:
        print(f"[ingest] ERREUR : {exc}", file=sys.stderr)
        raise

    # Critères d'acceptation : ~1 067 371 lignes attendues
    expected = 1_067_371
    if abs(result["rows_inserted"] - expected) > 100:
        print(
            f"[ingest] ⚠️ Volumétrie inattendue : "
            f"{result['rows_inserted']:,} vs {expected:,} attendu",
            file=sys.stderr,
        )
        return 2
    return 0


if __name__ == "__main__":
    sys.exit(main())
